
import openpyxl
from bs4 import BeautifulSoup

# Read the HTML file
with open("sk.html", "r", encoding="utf8") as file:
    html_data = file.read()

# Parse the HTML using BeautifulSoup
soup = BeautifulSoup(html_data, "html.parser")

# Find all div elements with the given attributes
div_elements = soup.find_all("div", attrs={"class": "s-card-container s-overflow-hidden aok-relative puis-wide-grid-style puis-wide-grid-style-t3 puis-expand-height puis-include-content-margin puis puis-v132n5e4faosf42v0eo3rf7vw9m s-latency-cf-section s-card-border"})

# Create lists to store the extracted information
product_names = []
product_prices = []
product_reviews = []

# Extract the desired information from each div element
for div in div_elements:
    # Find the product name
    name_span = div.find("span", class_="a-size-base-plus a-color-base a-text-normal")
    product_name = name_span.get_text(strip=True) if name_span else ""
    product_names.append(product_name)

    # Find the product price
    price_span = div.find("span", class_="a-price-whole")
    product_price = price_span.get_text(strip=True) if price_span else ""
    product_prices.append(product_price)

    # Find the product review
    review_span = div.find("span", class_="a-size-base s-underline-text")
    product_review = review_span.get_text(strip=True) if review_span else ""
    product_reviews.append(product_review)

# Create a new Excel file and write the extracted information
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers
worksheet["A1"] = "Product Name"
worksheet["B1"] = "Product Price"
worksheet["C1"] = "Product Review"

# Write the data
for i in range(len(product_names)):
    worksheet.cell(row=i+2, column=1, value=product_names[i])
    worksheet.cell(row=i+2, column=2, value=product_prices[i])
    worksheet.cell(row=i+2, column=3, value=product_reviews[i])

# Save the Excel file
workbook.save("Data.xlsx")
